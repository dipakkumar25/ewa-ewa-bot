# src/ewa_processor.py
"""
EWA processor for A1C using embedded Excel traffic-light matrix.

Pipeline:
- For each .docx report in DOCX_DIR:
    - parse system + report_date from filename
    - read overall status from first inline image on first page
    - open embedded Excel from /word/embeddings/oleObject*.bin
    - read KPI rows and latest colored cell
    - map fill RGB -> GREEN / YELLOW / RED / NA
- Save full history to HISTORY_CSV
- Train RandomForest model to predict overall_status from KPIs
"""

import os
import sys
import re
import io
from datetime import datetime
from typing import Dict, List, Tuple

# ensure project root is importable
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import zipfile

import joblib
import olefile
import openpyxl
import pandas as pd
from docx import Document
from PIL import Image
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import classification_report

from src.config import (
    DOCX_DIR,
    HISTORY_CSV,
    MODEL_FILE,
    SECTION_KEYWORDS,
    KPI_ALIASES,
    COLOR_MAP,
)


# ---------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------


def parse_filename(filename: str) -> Tuple[str, datetime]:
    """
    Expected filename, e.g.:
    A1C_21277797_850764463_2025-11-24_R_EWA.docx
    """
    base = os.path.basename(filename)
    system = base.split("_")[0]
    m = re.search(r"\d{4}-\d{2}-\d{2}", base)
    if not m:
        raise ValueError(f"Could not find date in filename: {filename}")
    dt = datetime.strptime(m.group(0), "%Y-%m-%d")
    return system, dt


def normalize_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "")).strip().lower()


def detect_color_from_rgb(r: int, g: int, b: int) -> str:
    """
    Very simple thresholds tuned for typical SAP traffic-light colors.
    """
    if r > 200 and g < 80 and b < 80:
        return "RED"
    if r > 200 and g > 200 and b < 80:
        return "YELLOW"
    if r < 80 and g > 200 and b < 80:
        return "GREEN"
    return "NA"


# ---------------------------------------------------------------------
# overall status from Word front-page icon
# ---------------------------------------------------------------------


def extract_overall_traffic_light(doc: Document) -> str:
    """
    Use the first inline image (traffic light) as overall status.
    If nothing can be read, 'NA'.
    """
    try:
        for ish in doc.inline_shapes:
            try:
                blob = ish.image.blob
                img = Image.open(io.BytesIO(blob)).convert("RGB")
                w, h = img.size
                r, g, b = img.getpixel((w // 2, h // 2))
                return detect_color_from_rgb(r, g, b)
            except Exception:
                continue
    except Exception:
        pass
    return "NA"


# ---------------------------------------------------------------------
# Excel embedded in the DOCX (.bin OLE -> xlsx -> openpyxl)
# ---------------------------------------------------------------------


def _extract_embedded_xlsx_bytes(docx_path: str) -> bytes:
    """
    Open the docx as a zip, locate /word/embeddings/oleObject*.bin,
    read using olefile and return the OPC 'Package' stream as bytes.
    This is effectively the embedded .xlsx file.
    """
    with zipfile.ZipFile(docx_path, "r") as zf:
        embedding_names = [
            name
            for name in zf.namelist()
            if name.startswith("word/embeddings/") and name.lower().endswith(".bin")
        ]
        if not embedding_names:
            raise RuntimeError(f"No embedded Excel found in {docx_path}")

        # Assume first embedding is the KPI matrix
        emb_name = embedding_names[0]
        bin_data = zf.read(emb_name)

    # Use olefile to read OLE container; 'Package' stream holds xlsx bytes
    with olefile.OleFileIO(io.BytesIO(bin_data)) as ole:
        if ole.exists("Package"):
            pkg = ole.openstream("Package")
            return pkg.read()
        # Fallback: sometimes workbook is in 'Workbook'
        if ole.exists("Workbook"):
            wb_stream = ole.openstream("Workbook")
            return wb_stream.read()
        raise RuntimeError("No 'Package' or 'Workbook' stream in embedded object")


def _open_embedded_workbook(docx_path: str) -> openpyxl.Workbook:
    """
    Return openpyxl workbook loaded from embedded Excel in a docx.
    """
    xlsx_bytes = _extract_embedded_xlsx_bytes(docx_path)
    bio = io.BytesIO(xlsx_bytes)
    wb = openpyxl.load_workbook(filename=bio, data_only=True)
    return wb


def _color_from_fill(fill) -> str:
    """
    Map openpyxl cell.fill to GREEN/YELLOW/RED/NA.
    """
    rgb = None

    try:
        if fill is None:
            return "NA"
        # for solid fills fgColor.rgb is 'FF00B050' etc
        if fill.fgColor is not None and fill.fgColor.type == "rgb":
            rgb = fill.fgColor.rgb  # ARGB
    except Exception:
        rgb = None

    if not rgb:
        return "NA"

    # rgb can be 'FF00B050' -> strip alpha
    rgb = rgb[-6:]
    try:
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
    except Exception:
        return "NA"

    return detect_color_from_rgb(r, g, b)


def extract_kpi_colors_from_excel(docx_path: str) -> Dict[str, str]:
    """
    Main logic:
    - open embedded Excel
    - assume first sheet contains KPI grid
    - first column = KPI names
    - later columns = weekly values with colored fills
    - for each KPI alias, find matching row and take *right-most* non-NA color.
    Returns {canonical_kpi_name: 'RED'/'YELLOW'/'GREEN'/'NA'}
    """
    wb = _open_embedded_workbook(docx_path)
    ws = wb.worksheets[0]  # assume first sheet

    # build mapping from row index to KPI canonical name
    row_to_kpi: Dict[int, str] = {}
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        text = normalize_text(str(cell.value) if cell.value is not None else "")
        for alias, canonical in KPI_ALIASES.items():
            if alias in text:
                row_to_kpi[cell.row] = canonical
                break

    # initialize result
    result: Dict[str, str] = {canon: "NA" for canon in SECTION_KEYWORDS}

    if not row_to_kpi:
        # nothing found – return NA for everything
        return result

    # for each KPI row, scan from right-most column back
    max_col = ws.max_column
    for row_idx, canonical in row_to_kpi.items():
        status = "NA"
        for col_idx in range(max_col, 1, -1):  # skip KPI name col=1
            cell = ws.cell(row=row_idx, column=col_idx)
            color = _color_from_fill(cell.fill)
            if color != "NA":
                status = color
                break
        result[canonical] = status

    return result


# ---------------------------------------------------------------------
# public functions used by Streamlit app
# ---------------------------------------------------------------------


def parse_single_report(path: str) -> Dict:
    """
    Parse one docx -> dict:
      system, report_date, overall_status, all KPI statuses
    """
    system, dt = parse_filename(path)
    doc = Document(path)

    rec: Dict[str, str] = {
        "system": system,
        "report_date": dt,
        "overall_status": extract_overall_traffic_light(doc),
    }

    try:
        kpi_colors = extract_kpi_colors_from_excel(path)
    except Exception as e:
        print(f"!! Excel extraction failed for {os.path.basename(path)}:", e)
        # fall back to NA for all KPIs
        kpi_colors = {canon: "NA" for canon in SECTION_KEYWORDS}

    rec.update(kpi_colors)

    # normalise: never None
    for k, v in list(rec.items()):
        if v is None:
            rec[k] = "NA"

    return rec


def build_history_from_folder() -> pd.DataFrame:
    """
    Parse all .docx in DOCX_DIR and build history CSV.
    """
    records: List[Dict] = []

    if not os.path.isdir(DOCX_DIR):
        raise RuntimeError(f"DOCX_DIR does not exist: {DOCX_DIR}")

    for fname in sorted(os.listdir(DOCX_DIR)):
        if not fname.lower().endswith(".docx"):
            continue
        full = os.path.join(DOCX_DIR, fname)
        print("Parsing:", fname)
        try:
            rec = parse_single_report(full)
            records.append(rec)
        except Exception as e:
            print("  !! Failed to parse", fname, "->", e)

    if not records:
        raise RuntimeError(f"No .docx reports found in {DOCX_DIR}")

    df = pd.DataFrame(records).sort_values("report_date").reset_index(drop=True)
    df.to_csv(HISTORY_CSV, index=False)
    print("✔ History saved to", HISTORY_CSV)
    return df


def encode_colors(df: pd.DataFrame) -> pd.DataFrame:
    enc = df.copy()
    for col in enc.columns:
        if col in ("system", "report_date"):
            continue
        enc[col] = enc[col].map(COLOR_MAP).fillna(-1).astype(int)
    return enc


def train_baseline_model(history_df: pd.DataFrame):
    enc = encode_colors(history_df)
    feature_cols = [
        c for c in enc.columns if c not in ("system", "report_date", "overall_status")
    ]
    X = enc[feature_cols].values
    y = enc["overall_status"].values

    split = max(1, int(len(enc) * 0.75))
    X_train, X_test = X[:split], X[split:]
    y_train, y_test = y[:split], y[split:]

    clf = RandomForestClassifier(
        n_estimators=250,
        random_state=42,
        class_weight="balanced",
    )
    clf.fit(X_train, y_train)

    if len(X_test) > 0:
        print("=== Baseline model performance ===")
        print(classification_report(y_test, clf.predict(X_test)))

    os.makedirs(os.path.dirname(MODEL_FILE), exist_ok=True)
    joblib.dump((clf, feature_cols), MODEL_FILE)
    print("✔ Model saved to", MODEL_FILE)


def load_model():
    return joblib.load(MODEL_FILE)


def predict_overall_from_sections(model, feature_cols, row: pd.Series) -> str:
    enc = encode_colors(pd.DataFrame([row]))
    pred = int(model.predict(enc[feature_cols])[0])
    inv = {v: k for k, v in COLOR_MAP.items()}
    return inv.get(pred, "NA")


def compute_deviation(history_df: pd.DataFrame, new_row: pd.Series) -> pd.DataFrame:
    last = history_df.iloc[-1]
    changes = []
    for col in history_df.columns:
        if col in ("system", "report_date"):
            continue
        if str(last[col]) != str(new_row[col]):
            changes.append({"metric": col, "previous": last[col], "new": new_row[col]})
    return pd.DataFrame(changes)


def score_risk(history_df: pd.DataFrame, new_row: pd.Series):
    last = history_df.iloc[-1]
    score = 0.0
    details = []

    for col in history_df.columns:
        if col in ("system", "report_date"):
            continue
        old = str(last[col])
        new = str(new_row[col])
        old_n = COLOR_MAP.get(old, -1)
        new_n = COLOR_MAP.get(new, -1)
        if new_n > old_n and new_n >= 0:
            inc = 2.0 if (old == "GREEN" and new == "RED") else 1.0
            score += inc
            details.append((col, old, new, inc))

    if score >= 3:
        level = "HIGH"
    elif score >= 1:
        level = "MEDIUM"
    else:
        level = "LOW"
    return level, score, details


# ---------------------------------------------------------------------
# CLI entrypoint
# ---------------------------------------------------------------------

if __name__ == "__main__":
    df_hist = build_history_from_folder()
    train_baseline_model(df_hist)
